import pandas as pd
from google.cloud import bigquery
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
from datetime import datetime
from app.utils.config import PROJECT_ID, TARGET_DATASET


class ReportGenerator:
    """
    Generate formatted Excel reports from dq_watchtower_results data.
    """

    def __init__(self):
        self.bq_client = bigquery.Client(project=PROJECT_ID)
        self.dataset = TARGET_DATASET
        self.report_date = datetime.now().strftime("%Y-%m-%d")

    def fetch_dq_results(self):
        """
        Fetch DQ check results from dq_watchtower_results table.
        Filters to show only the latest execution date and latest execution per rule to prevent data duplication.
        """
        try:
            query = f"""
            WITH latest_rules AS (
                SELECT 
                    execution_ts,
                    run_id,
                    table_name,
                    column_name,
                    rule_name,
                    total_records,
                    passed_records,
                    failed_records,
                    pass_percentage,
                    execution_time_ms,
                    execution_status,
                    dq_status,
                    ROW_NUMBER() OVER (PARTITION BY table_name, column_name, rule_name ORDER BY execution_ts DESC) AS rn
                FROM `{PROJECT_ID}.{self.dataset}.dq_watchtower_results`
                WHERE DATE(execution_ts) = (SELECT MAX(DATE(execution_ts)) FROM `{PROJECT_ID}.{self.dataset}.dq_watchtower_results`)
            )
            SELECT 
                execution_ts,
                run_id,
                table_name,
                column_name,
                rule_name,
                total_records,
                passed_records,
                failed_records,
                pass_percentage,
                execution_time_ms,
                execution_status,
                dq_status
            FROM latest_rules
            WHERE rn = 1
            ORDER BY execution_ts DESC
            LIMIT 500
            """
            
            query_job = self.bq_client.query(query)
            results = query_job.result()
            
            print(f"✅ Fetched {results.total_rows} rows from dq_watchtower_results")
            
            # Convert to DataFrame
            data = []
            for row in results:
                data.append({
                    "Timestamp": str(row.execution_ts),
                    "Run ID": row.run_id if row.run_id else "",
                    "Table Name": row.table_name,
                    "Column Name": row.column_name if row.column_name else "",
                    "Rule Name": row.rule_name,
                    "Total Records": row.total_records if row.total_records else 0,
                    "Passed": row.passed_records if row.passed_records else 0,
                    "Failed": row.failed_records if row.failed_records else 0,
                    "Pass %": f"{row.pass_percentage:.1f}%" if row.pass_percentage else "0.0%",
                    "Execution Time (ms)": row.execution_time_ms if row.execution_time_ms else 0,
                    "Execution Status": row.execution_status if row.execution_status else "",
                    "Status": row.dq_status if row.dq_status else "UNKNOWN"
                })
            
            return pd.DataFrame(data)
        
        except Exception as e:
            print(f"ERROR fetching DQ results: {str(e)}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def fetch_summary_data(self):
        """
        Fetch summary statistics using the latest DQ results per unique check.
        Includes Project ID and Dataset Name for all tables.
        """
        try:
            query = f"""
            WITH latest_dq AS (
                SELECT *
                FROM (
                    SELECT 
                        table_name,
                        column_name,
                        rule_name,
                        total_records,
                        passed_records,
                        failed_records,
                        dq_status,
                        execution_ts,
                        ROW_NUMBER() OVER (PARTITION BY table_name, column_name, rule_name ORDER BY execution_ts DESC) AS rn
                    FROM `{PROJECT_ID}.{self.dataset}.dq_watchtower_results`
                )
                WHERE rn = 1
            ),
            table_summary AS (
                SELECT
                    table_name,
                    COUNT(*) AS total_checks,
                    SUM(CASE WHEN dq_status = 'PASS' THEN 1 ELSE 0 END) AS passed_checks,
                    SUM(CASE WHEN dq_status = 'FAIL' THEN 1 ELSE 0 END) AS failed_checks,
                    ROUND(100.0 * SUM(CASE WHEN dq_status = 'PASS' THEN 1 ELSE 0 END) / COUNT(*), 2) AS check_health_pct,
                    MAX(total_records) AS total_records,
                    SUM(passed_records) AS passed_records,
                    ROUND(100.0 * SUM(passed_records) / SUM(total_records), 2) AS record_pass_pct
                FROM latest_dq
                GROUP BY table_name
            )
            SELECT *
            FROM table_summary
            ORDER BY table_name
            """
            
            query_job = self.bq_client.query(query)
            results = query_job.result()
            
            print(f"✅ Fetched summary data for {results.total_rows} tables")
            
            # Convert to DataFrame
            data = []
            for row in results:
                data.append({
                    "Project ID": PROJECT_ID,
                    "Dataset Name": self.dataset,
                    "Table Name": row.table_name,
                    "Total Checks": row.total_checks,
                    "Passed Checks": row.passed_checks,
                    "Failed Checks": row.failed_checks,
                    "Check Health %": row.check_health_pct,
                    "Total Records": row.total_records if row.total_records else 0,
                    "Passed Records": row.passed_records if row.passed_records else 0,
                    "Record Pass %": row.record_pass_pct if row.record_pass_pct else 0
                })
            
            return pd.DataFrame(data)
        
        except Exception as e:
            print(f"ERROR fetching summary data: {str(e)}")
            import traceback
            traceback.print_exc()
            return pd.DataFrame()

    def generate_summary_stats(self, summary_df):
        """
        Generate overall summary statistics from summary data.
        """
        if summary_df.empty:
            return {
                "Total Checks": 0,
                "Passed": 0,
                "Failed": 0,
                "Pass Rate": "0%"
            }
        
        total = summary_df["Total Checks"].sum()
        passed = summary_df["Passed Checks"].sum()
        failed = summary_df["Failed Checks"].sum()
        pass_rate = f"{(passed/total*100):.1f}%" if total > 0 else "0%"
        
        return {
            "Total Checks": int(total),
            "Passed": int(passed),
            "Failed": int(failed),
            "Pass Rate": pass_rate
        }

    def create_excel_report(self, df):
        """
        Create formatted Excel report with summary and detailed tables.
        Returns BytesIO object (in-memory Excel file).
        """
        # Fetch summary data using the new query
        summary_df = self.fetch_summary_data()
        
        wb = Workbook()
        
        # Define styles - Light & Soft Pastel Colors
        header_fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
        header_font = Font(bold=True, color="2C5282", size=12)
        
        good_fill = PatternFill(start_color="F0F8E8", end_color="F0F8E8", fill_type="solid")
        good_font = Font(color="2D5016", bold=True)
        
        failed_fill = PatternFill(start_color="FDF0F0", end_color="FDF0F0", fill_type="solid")
        failed_font = Font(color="742C2C", bold=True)
        
        warning_fill = PatternFill(start_color="FEF9E8", end_color="FEF9E8", fill_type="solid")
        warning_font = Font(color="7A6928", bold=True)
        
        border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin")
        )
        
        center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
        left_align = Alignment(horizontal="left", vertical="center", wrap_text=True)
        
        # ========== SUMMARY SHEET ==========
        ws_summary = wb.active
        ws_summary.title = "Summary"
        
        # Title
        ws_summary["A1"] = f"DQ Health Report - {self.report_date}"
        ws_summary["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_summary["A1"].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws_summary.merge_cells("A1:I1")
        ws_summary["A1"].alignment = center_align
        ws_summary.row_dimensions[1].height = 25
        
        # ========== PROJECT INFO (TOP) ==========
        row = 3
        ws_summary[f"A{row}"] = "Project ID"
        ws_summary[f"B{row}"] = PROJECT_ID
        ws_summary[f"A{row}"].font = Font(bold=True, size=11)
        ws_summary[f"B{row}"].font = Font(size=11)
        ws_summary[f"A{row}"].border = border
        ws_summary[f"B{row}"].border = border
        ws_summary[f"A{row}"].alignment = left_align
        ws_summary[f"B{row}"].alignment = left_align
        
        row += 1
        ws_summary[f"A{row}"] = "Dataset Name"
        ws_summary[f"B{row}"] = self.dataset
        ws_summary[f"A{row}"].font = Font(bold=True, size=11)
        ws_summary[f"B{row}"].font = Font(size=11)
        ws_summary[f"A{row}"].border = border
        ws_summary[f"B{row}"].border = border
        ws_summary[f"A{row}"].alignment = left_align
        ws_summary[f"B{row}"].alignment = left_align
        
        row += 1
        ws_summary[f"A{row}"] = "Report Generated"
        ws_summary[f"B{row}"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S IST")
        ws_summary[f"A{row}"].font = Font(bold=True, size=11)
        ws_summary[f"B{row}"].font = Font(size=11)
        ws_summary[f"A{row}"].border = border
        ws_summary[f"B{row}"].border = border
        ws_summary[f"A{row}"].alignment = left_align
        ws_summary[f"B{row}"].alignment = left_align
        
        # ========== SECTION 1: EXECUTIVE SUMMARY ==========
        row += 2
        ws_summary[f"A{row}"] = "Executive Summary"
        ws_summary[f"A{row}"].font = Font(bold=True, size=11, color="2C5282")
        ws_summary[f"A{row}"].fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
        ws_summary.merge_cells(f"A{row}:F{row}")
        ws_summary[f"A{row}"].alignment = left_align
        ws_summary.row_dimensions[row].height = 16
        
        row += 1
        
        # Prepare Executive Summary Data
        total_tables = summary_df["Table Name"].nunique() if not summary_df.empty else 0
        
        # Calculate correct DQ Health % based on total passed/failed checks across all tables
        if not summary_df.empty:
            total_checks_all = int(summary_df["Total Checks"].sum())
            passed_checks_all = int(summary_df["Passed Checks"].sum())
            failed_checks_all = int(summary_df["Failed Checks"].sum())
            dq_health = f"{(100.0 * passed_checks_all / total_checks_all):.1f}%" if total_checks_all > 0 else "0.0%"
        else:
            total_checks_all = 0
            passed_checks_all = 0
            failed_checks_all = 0
            dq_health = "0.0%"
        
        total_rules = str(total_checks_all)
        passed_rules = str(passed_checks_all)
        failed_rules = str(failed_checks_all)
        table_count = str(total_tables)
        
        # Count tables with failed checks
        impacted_tables = len(summary_df[summary_df["Failed Checks"] > 0]) if not summary_df.empty else 0
        
        # Horizontal Headers (Metrics as columns)
        headers = ["Total Rules Evaluated", "Rules Passed", "Rules Failed", "Tables Monitored", "Tables Impacted", "Overall DQ Health"]
        for col, header in enumerate(headers, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        
        # Horizontal Values (single row)
        values = [total_rules, passed_rules, failed_rules, table_count, impacted_tables, dq_health]
        for col, value in enumerate(values, 1):
            cell = ws_summary.cell(row=row, column=col)
            cell.value = value
            cell.border = border
            cell.alignment = center_align
        
        row += 1
        
        # ========== SECTION 2: HEALTH BY TABLE ==========
        row += 1
        ws_summary[f"A{row}"] = "Health by Table"
        ws_summary[f"A{row}"].font = Font(bold=True, size=11, color="2C5282")
        ws_summary[f"A{row}"].fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
        ws_summary.merge_cells(f"A{row}:I{row}")
        ws_summary[f"A{row}"].alignment = left_align
        ws_summary.row_dimensions[row].height = 16
        
        row += 1
        
        if not summary_df.empty:
            # Headers
            headers = ["Project ID", "Dataset Name", "Table Name", "Total Checks", "Passed Checks", "Failed Checks", "Total Records", "Passed Records", "Record Pass %"]
            for col, header in enumerate(headers, 1):
                cell = ws_summary.cell(row=row, column=col)
                cell.value = header
                cell.fill = header_fill
                cell.font = header_font
                cell.border = border
                cell.alignment = center_align
            
            row += 1
            
            # Data
            for _, r in summary_df.iterrows():
                ws_summary.cell(row=row, column=1).value = r["Project ID"]
                ws_summary.cell(row=row, column=1).alignment = center_align
                ws_summary.cell(row=row, column=1).border = border
                
                ws_summary.cell(row=row, column=2).value = r["Dataset Name"]
                ws_summary.cell(row=row, column=2).alignment = center_align
                ws_summary.cell(row=row, column=2).border = border
                
                ws_summary.cell(row=row, column=3).value = r["Table Name"]
                ws_summary.cell(row=row, column=3).alignment = center_align
                ws_summary.cell(row=row, column=3).border = border
                
                ws_summary.cell(row=row, column=4).value = int(r["Total Checks"])
                ws_summary.cell(row=row, column=4).alignment = center_align
                ws_summary.cell(row=row, column=4).border = border
                
                ws_summary.cell(row=row, column=5).value = int(r["Passed Checks"])
                ws_summary.cell(row=row, column=5).alignment = center_align
                ws_summary.cell(row=row, column=5).border = border
                
                ws_summary.cell(row=row, column=6).value = int(r["Failed Checks"])
                ws_summary.cell(row=row, column=6).alignment = center_align
                ws_summary.cell(row=row, column=6).border = border
                
                ws_summary.cell(row=row, column=7).value = int(r["Total Records"])
                ws_summary.cell(row=row, column=7).alignment = center_align
                ws_summary.cell(row=row, column=7).border = border
                
                ws_summary.cell(row=row, column=8).value = int(r["Passed Records"])
                ws_summary.cell(row=row, column=8).alignment = center_align
                ws_summary.cell(row=row, column=8).border = border
                
                ws_summary.cell(row=row, column=9).value = f"{r['Record Pass %']:.1f}%"
                ws_summary.cell(row=row, column=9).alignment = center_align
                ws_summary.cell(row=row, column=9).border = border
                
                row += 1
        
        # ========== SECTION 3: DQ OUTAGES REQUIRING ATTENTION ==========
        row += 2
        ws_summary[f"A{row}"] = "DQ Outages Requiring Attention"
        ws_summary[f"A{row}"].font = Font(bold=True, size=11, color="2C5282")
        ws_summary[f"A{row}"].fill = PatternFill(start_color="E8F0F8", end_color="E8F0F8", fill_type="solid")
        ws_summary.merge_cells(f"A{row}:F{row}")
        ws_summary[f"A{row}"].alignment = left_align
        ws_summary.row_dimensions[row].height = 16
        
        row += 1
        
        if not summary_df.empty:
            # Sort by failed checks descending and get top 10
            burnout_df = summary_df[summary_df["Failed Checks"] > 0].sort_values("Failed Checks", ascending=False).head(10)
            
            if not burnout_df.empty:
                # Headers
                headers = ["Project ID", "Dataset Name", "Table Name", "Total Rules", "Failed Rules", "Health %"]
                for col, header in enumerate(headers, 1):
                    cell = ws_summary.cell(row=row, column=col)
                    cell.value = header
                    cell.fill = header_fill
                    cell.font = header_font
                    cell.border = border
                    cell.alignment = center_align
                
                row += 1
                
                # Data
                for _, r in burnout_df.iterrows():
                    ws_summary.cell(row=row, column=1).value = r["Project ID"]
                    ws_summary.cell(row=row, column=1).alignment = center_align
                    ws_summary.cell(row=row, column=1).border = border
                    
                    ws_summary.cell(row=row, column=2).value = r["Dataset Name"]
                    ws_summary.cell(row=row, column=2).alignment = center_align
                    ws_summary.cell(row=row, column=2).border = border
                    
                    ws_summary.cell(row=row, column=3).value = r["Table Name"]
                    ws_summary.cell(row=row, column=3).alignment = center_align
                    ws_summary.cell(row=row, column=3).border = border
                    
                    ws_summary.cell(row=row, column=4).value = int(r["Total Checks"])
                    ws_summary.cell(row=row, column=4).alignment = center_align
                    ws_summary.cell(row=row, column=4).border = border
                    
                    ws_summary.cell(row=row, column=5).value = int(r["Failed Checks"])
                    ws_summary.cell(row=row, column=5).alignment = center_align
                    ws_summary.cell(row=row, column=5).border = border
                    
                    ws_summary.cell(row=row, column=6).value = f"{r['Check Health %']:.1f}%"
                    ws_summary.cell(row=row, column=6).alignment = center_align
                    ws_summary.cell(row=row, column=6).border = border
                    
                    row += 1
            else:
                ws_summary[f"A{row}"] = "All tables healthy - No issues"
                ws_summary[f"A{row}"].fill = good_fill
                ws_summary[f"A{row}"].font = good_font
                ws_summary.merge_cells(f"A{row}:F{row}")
                ws_summary[f"A{row}"].alignment = center_align
                ws_summary[f"A{row}"].border = border
        
        # Adjust column widths
        ws_summary.column_dimensions["A"].width = 18
        ws_summary.column_dimensions["B"].width = 18
        ws_summary.column_dimensions["C"].width = 20
        ws_summary.column_dimensions["D"].width = 20
        ws_summary.column_dimensions["E"].width = 14
        ws_summary.column_dimensions["F"].width = 14
        ws_summary.column_dimensions["G"].width = 14
        ws_summary.column_dimensions["H"].width = 14
        ws_summary.column_dimensions["I"].width = 14
        ws_summary.column_dimensions["J"].width = 14
        ws_summary.column_dimensions["K"].width = 18
        
        # ========== DETAILED SHEET ==========
        ws_detail = wb.create_sheet("Detailed Results")
        
        # Title
        ws_detail["A1"] = f"📋 Detailed DQ Check Results - {self.report_date}"
        ws_detail["A1"].font = Font(bold=True, size=14, color="FFFFFF")
        ws_detail["A1"].fill = PatternFill(start_color="203864", end_color="203864", fill_type="solid")
        ws_detail.merge_cells("A1:L1")
        ws_detail["A1"].alignment = center_align
        ws_detail.row_dimensions[1].height = 25
        
        # Headers - ALL COLUMNS
        headers = [
            "Execution Timestamp",
            "Run ID",
            "Table Name",
            "Column Name",
            "Rule Name",
            "Total Records",
            "Passed Records",
            "Failed Records",
            "Pass %",
            "Execution Time (ms)",
            "Execution Status",
            "DQ Status"
        ]
        for col, header in enumerate(headers, 1):
            cell = ws_detail.cell(row=3, column=col)
            cell.value = header
            cell.fill = header_fill
            cell.font = header_font
            cell.border = border
            cell.alignment = center_align
        
        # Data
        row = 4
        for _, r in df.iterrows():
            # Execution Timestamp
            ws_detail.cell(row=row, column=1).value = r["Timestamp"]
            ws_detail.cell(row=row, column=1).alignment = left_align
            ws_detail.cell(row=row, column=1).border = border
            
            # Run ID
            ws_detail.cell(row=row, column=2).value = r.get("Run ID", "")
            ws_detail.cell(row=row, column=2).alignment = left_align
            ws_detail.cell(row=row, column=2).border = border
            
            # Table Name
            ws_detail.cell(row=row, column=3).value = r["Table Name"]
            ws_detail.cell(row=row, column=3).alignment = left_align
            ws_detail.cell(row=row, column=3).border = border
            
            # Column Name
            ws_detail.cell(row=row, column=4).value = r["Column Name"]
            ws_detail.cell(row=row, column=4).alignment = left_align
            ws_detail.cell(row=row, column=4).border = border
            
            # Rule Name
            ws_detail.cell(row=row, column=5).value = r["Rule Name"]
            ws_detail.cell(row=row, column=5).alignment = left_align
            ws_detail.cell(row=row, column=5).border = border
            
            # Total Records
            total_records_cell = ws_detail.cell(row=row, column=6)
            total_records_cell.value = int(r["Total Records"]) if r["Total Records"] else 0
            total_records_cell.alignment = center_align
            total_records_cell.border = border
            
            # Passed Records - GREEN
            passed_records_cell = ws_detail.cell(row=row, column=7)
            passed_records_cell.value = int(r["Passed"]) if r["Passed"] else 0
            passed_records_cell.fill = good_fill
            passed_records_cell.font = good_font
            passed_records_cell.alignment = center_align
            passed_records_cell.border = border
            
            # Failed Records - RED (if any)
            failed_records_cell = ws_detail.cell(row=row, column=8)
            failed_records_cell.value = int(r["Failed"]) if r["Failed"] else 0
            if int(r["Failed"]) if r["Failed"] else 0 > 0:
                failed_records_cell.fill = failed_fill
                failed_records_cell.font = failed_font
            else:
                failed_records_cell.fill = good_fill
                failed_records_cell.font = good_font
            failed_records_cell.alignment = center_align
            failed_records_cell.border = border
            
            # Pass % - COLOR CODED
            pass_pct_cell = ws_detail.cell(row=row, column=9)
            pass_pct_cell.value = r["Pass %"]
            try:
                pass_pct_value = float(str(r["Pass %"]).rstrip("%"))
                if pass_pct_value >= 95:
                    pass_pct_cell.fill = good_fill
                    pass_pct_cell.font = good_font
                elif pass_pct_value >= 80:
                    pass_pct_cell.fill = warning_fill
                    pass_pct_cell.font = warning_font
                else:
                    pass_pct_cell.fill = failed_fill
                    pass_pct_cell.font = failed_font
            except:
                pass
            pass_pct_cell.alignment = center_align
            pass_pct_cell.border = border
            
            # Execution Time (ms)
            exec_time_cell = ws_detail.cell(row=row, column=10)
            exec_time_cell.value = r.get("Execution Time (ms)", "")
            exec_time_cell.alignment = center_align
            exec_time_cell.border = border
            
            # Execution Status
            exec_status_cell = ws_detail.cell(row=row, column=11)
            exec_status_cell.value = r["Execution Status"]
            exec_status_cell.alignment = center_align
            exec_status_cell.border = border
            if str(r["Execution Status"]).lower() == "success":
                exec_status_cell.fill = good_fill
                exec_status_cell.font = good_font
            else:
                exec_status_cell.fill = failed_fill
                exec_status_cell.font = failed_font
            
            # DQ Status - WITH COLOR CODING
            dq_status_cell = ws_detail.cell(row=row, column=12)
            dq_status_cell.value = r["Status"]
            dq_status_cell.alignment = center_align
            dq_status_cell.border = border
            
            # Color code: GREEN for PASS, RED for FAIL
            status_upper = str(r["Status"]).upper()
            if status_upper == "PASS":
                dq_status_cell.fill = good_fill
                dq_status_cell.font = good_font
            else:
                dq_status_cell.fill = failed_fill
                dq_status_cell.font = failed_font
            
            row += 1
        
        # Adjust column widths
        ws_detail.column_dimensions["A"].width = 22
        ws_detail.column_dimensions["B"].width = 38
        ws_detail.column_dimensions["C"].width = 18
        ws_detail.column_dimensions["D"].width = 18
        ws_detail.column_dimensions["E"].width = 18
        ws_detail.column_dimensions["F"].width = 14
        ws_detail.column_dimensions["G"].width = 14
        ws_detail.column_dimensions["H"].width = 14
        ws_detail.column_dimensions["I"].width = 10
        ws_detail.column_dimensions["J"].width = 14
        ws_detail.column_dimensions["K"].width = 14
        ws_detail.column_dimensions["L"].width = 12
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        return output

    def generate_report(self):
        """
        Main function: Fetch data, generate Excel report.
        Returns BytesIO object.
        """
        print("📊 Fetching DQ check results...")
        df = self.fetch_dq_results()
        
        if df.empty:
            print("⚠️ No DQ check results found, creating empty report...")
            # Create empty report anyway
            wb = Workbook()
            ws = wb.active
            ws.title = "No Data"
            ws["A1"] = "No DQ check results found for this period"
            
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            return output
        
        print(f"📝 Generating Excel report with {len(df)} records...")
        excel_file = self.create_excel_report(df)
        
        print("✅ Excel report generated successfully")
        return excel_file
